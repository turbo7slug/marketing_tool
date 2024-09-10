import os
from flask import Flask, request, jsonify, send_file
import cv2
import numpy as np
import pytesseract
from pdf2image import convert_from_path
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as ExcelImage
import uuid
from flask_cors import CORS
from dotenv import load_dotenv
import re 

# Load environment variables from .env file
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
CORS(app)

UPLOAD_FOLDER = os.getenv('UPLOAD_FOLDER', 'uploads')

# Ensure the uploads directory exists
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Function to detect product regions in the image
def detect_products(image):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edged = cv2.Canny(blurred, 50, 150)

    contours, _ = cv2.findContours(edged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    product_regions = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        if w > 100 and h > 100:
            product_regions.append((x, y, w, h))

    return sorted(product_regions, key=lambda r: r[1])

# Function to clean extracted text
def clean_text(text):
    text = re.sub(r'[^\x20-\x7E]', '', text)
    text = text.replace('\n', ' ').strip()
    return text

# Function to extract text from an image region using OCR
def extract_text(image, region):
    x, y, w, h = region
    product_image = image[y:y+h, x:x+w]

    gray = cv2.cvtColor(product_image, cv2.COLOR_BGR2GRAY)
    _, binary_image = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)

    config = "--oem 3 --psm 6"
    text = pytesseract.image_to_string(binary_image, config=config)
    return product_image, clean_text(text)

# Function to append data to the Excel sheet
def append_products_to_excel(product_details, excel_path):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.cell(row=1, column=1, value="Image")
        ws.cell(row=1, column=2, value="Description")

    last_row = ws.max_row

    for i, (product_image, description) in enumerate(product_details, start=last_row + 1):
        img_path = f"{UPLOAD_FOLDER}/product_{i-1}.png"
        cv2.imwrite(img_path, product_image)

        img = ExcelImage(img_path)
        img.anchor = f'A{i}'
        img.width = 100
        img.height = 100
        ws.add_image(img)

        ws.cell(row=i, column=2, value=description)

    wb.save(excel_path)

    for i in range(last_row + 1, last_row + len(product_details) + 1):
        os.remove(f"{UPLOAD_FOLDER}/product_{i-1}.png")

# Function to process PDF and extract product details
def process_pdf(pdf_path, excel_path, append=False):
    images = convert_from_path(pdf_path)

    all_product_details = []

    for page_number, image in enumerate(images, start=1):
        image_cv = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
        product_regions = detect_products(image_cv)
        product_details = [extract_text(image_cv, region) for region in product_regions]
        all_product_details.extend(product_details)

    append_products_to_excel(all_product_details, excel_path)

@app.route('/upload', methods=['POST'])
def upload_pdf():
    try:
        # Check if PDF is present
        if 'pdf' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        # Save the uploaded PDF
        pdf_file = request.files['pdf']
        pdf_filename = str(uuid.uuid4()) + ".pdf"
        pdf_path = os.path.join(UPLOAD_FOLDER, pdf_filename)
        pdf_file.save(pdf_path)

        # Handle appending to an existing Excel file
        append_to_existing = request.form.get('append') == 'true'
        if append_to_existing:
            if 'excel' not in request.files:
                return jsonify({"error": "Please upload an Excel file to append"}), 400
            excel_file = request.files['excel']
            excel_filename = excel_file.filename
            excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)
            excel_file.save(excel_path)
        else:
            excel_filename = str(uuid.uuid4()) + ".xlsx"
            excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)

        # Process the PDF and append or create Excel
        process_pdf(pdf_path, excel_path, append_to_existing)

        # Send the file for download
        return send_file(excel_path, as_attachment=True)

    except Exception as e:
        # Log the error for debugging purposes
        print(f"An error occurred: {str(e)}")
        return jsonify({"error": f"An error occurred: {str(e)}"}), 500

    finally:
        # Clean up
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        if not append_to_existing and os.path.exists(excel_path):
            os.remove(excel_path)

if __name__ == '__main__':
    app.run(debug=os.getenv('FLASK_DEBUG', 'False') == 'True', host=os.getenv('FLASK_HOST', '0.0.0.0'), port=int(os.getenv('FLASK_PORT', 5000)))
